import openpyxl

# Otwórz plik Excela
workbook = openpyxl.load_workbook('mszyce.xlsx')

# Wybierz arkusz
worksheet = workbook['Mieszkanie-Sprzedaz']

# Iteruj przez wszystkie wiersze
for row in worksheet.iter_rows():

    # Iteruj przez wszystkie komórki w wierszu
    for cell in row:

        # Sprawdź, czy wartość komórki to 'listingeditwindow_btn_save'
        if cell.value == 'listingeditwindow_btn_save':

            # Dodaj nowy wiersz pod wierszem z wartością 'listingeditwindow_btn_save'
            new_row = [cell.value for cell in row]
            worksheet.insert_rows(cell.row + 1)
            next_row = worksheet.cell(row=cell.row + 1, column=1)
            for idx, val in enumerate(new_row):
                next_row.offset(column=idx, row=0).value = val

# Zapisz zmiany w pliku Excela
workbook.save('wysylka-priorytetem-do-twojej-pizdy.xlsx')
